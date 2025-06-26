#!/usr/bin/env python3
"""
GST Intelligence Platform - Complete Main Application 
Version: 2.0.0 - All Features Integrated + Fixes Applied
"""

import os
import re
import asyncio
import asyncpg
import hashlib
import secrets
import logging
import json
import httpx
import fnmatch
import csv
import time
from datetime import datetime, timedelta
from functools import wraps
from decimal import Decimal
from collections import defaultdict
from typing import Optional, Dict, List, Any, Union
from io import BytesIO, StringIO

# FastAPI imports
from fastapi import FastAPI, Request, Form, Depends, HTTPException, status, Response, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles

# External libraries
from dotenv import load_dotenv
from pydantic import BaseModel, validator, Field

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Try importing optional dependencies with error handling
try:
    from weasyprint import HTML
    HAS_WEASYPRINT = True
    logger.info("✅ WeasyPrint available - PDF generation enabled")
except (ImportError, OSError, Exception) as e:
    logger.warning(f"⚠️ WeasyPrint not available - PDF generation disabled: {e}")
    HAS_WEASYPRINT = False
    HTML = None

# Try importing Redis
try:
    import redis.asyncio as redis
    HAS_REDIS = True
except ImportError:
    logger.warning("⚠️ Redis not available - using memory cache only")
    HAS_REDIS = False
    redis = None

# Local imports
try:
    from razorpay_lending import RazorpayLendingClient, LoanManager, LoanConfig, LoanStatus
    HAS_LOAN_MANAGEMENT = True
    logger.info("✅ Loan management available")
except ImportError:
    logger.warning("ℹ️ Loan management not available")
    HAS_LOAN_MANAGEMENT = False
    LoanConfig = None
    LoanStatus = None

# Try importing AI features
try:
    from anthro_ai import AnthropicClient
    HAS_AI_FEATURES = True
    logger.info("✅ AI features available")
except ImportError:
    logger.warning("ℹ️ AI features not available")
    HAS_AI_FEATURES = False
    AnthropicClient = None

# Import config
try:
    import config
    logger.info("✅ Configuration loaded")
except ImportError:
    logger.error("❌ Configuration not found - creating minimal config")
    class Config:
        POSTGRES_DSN = os.getenv("POSTGRES_DSN")
        SECRET_KEY = os.getenv("SECRET_KEY", "change-this-secret-key")
        RAPIDAPI_KEY = os.getenv("RAPIDAPI_KEY")
        RAPIDAPI_HOST = os.getenv("RAPIDAPI_HOST", "gst-return-status.p.rapidapi.com")
        REDIS_URL = os.getenv("REDIS_URL")
        ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
        RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID")
        RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET")
        RAZORPAY_ENVIRONMENT = os.getenv("RAZORPAY_ENVIRONMENT", "sandbox")
        ADMIN_USERS = os.getenv("ADMIN_USERS", "")
        RATE_LIMIT_REQUESTS = int(os.getenv("API_RATE_LIMIT", "100"))
        RATE_LIMIT_WINDOW = int(os.getenv("API_RATE_WINDOW", "3600"))
        is_production = os.getenv("ENVIRONMENT", "development").lower() == "production"
        SESSION_DURATION = timedelta(days=int(os.getenv("SESSION_DURATION_DAYS", "7")))
    config = Config()

# Import database
from database import DatabaseManager

# =============================================================================
# PYDANTIC MODELS
# =============================================================================

class UserSignupRequest(BaseModel):
    mobile: str = Field(..., min_length=10, max_length=10)
    password: str = Field(..., min_length=6, max_length=128)
    confirm_password: str = Field(..., min_length=6, max_length=128)
    
    @validator('mobile')
    def validate_mobile(cls, v):
        if not re.match(r'^[6-9]\d{9}$', v):
            raise ValueError('Invalid mobile number format')
        return v
    
    @validator('confirm_password')
    def passwords_match(cls, v, values):
        if 'password' in values and v != values['password']:
            raise ValueError('Passwords do not match')
        return v

class UserLoginRequest(BaseModel):
    mobile: str = Field(..., min_length=10, max_length=10)
    password: str = Field(..., min_length=1, max_length=128)

class GSTINSearchRequest(BaseModel):
    gstin: str = Field(..., min_length=15, max_length=15)
    
    @validator('gstin')
    def validate_gstin(cls, v):
        v = v.upper().strip()
        if not re.match(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$', v):
            raise ValueError('Invalid GSTIN format')
        return v

class BatchSearchRequest(BaseModel):
    gstins: List[str] = Field(..., min_items=1, max_items=10)
    
    @validator('gstins')
    def validate_gstins(cls, v):
        validated = []
        for gstin in v:
            gstin = gstin.upper().strip()
            if not re.match(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$', gstin):
                raise ValueError(f'Invalid GSTIN format: {gstin}')
            validated.append(gstin)
        return validated

class ChangePasswordRequest(BaseModel):
    currentPassword: str = Field(..., min_length=1)
    newPassword: str = Field(..., min_length=6)

class UserPreferencesRequest(BaseModel):
    theme: Optional[str] = Field(None, pattern=r'^(light|dark)$')
    notifications: Optional[bool] = None
    analytics: Optional[bool] = None
    auto_export: Optional[bool] = None

class UserProfileRequest(BaseModel):
    display_name: Optional[str] = Field(None, max_length=100)
    company: Optional[str] = Field(None, max_length=200)
    email: Optional[str] = Field(None, max_length=255)
    designation: Optional[str] = Field(None, max_length=100)
    
    @validator('email')
    def validate_email(cls, v):
        if v and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', v):
            raise ValueError('Invalid email format')
        return v

class ExportRequest(BaseModel):
    format: str = Field(..., pattern=r'^(csv|excel|pdf)$')
    type: str = Field(..., pattern=r'^(history|analytics|all)$')
    date_range: Optional[str] = Field(None, pattern=r'^(7days|30days|90days|all)$')

class ErrorLogRequest(BaseModel):
    type: str
    message: str
    stack: Optional[str] = None
    url: Optional[str] = None
    userAgent: Optional[str] = None
    timestamp: Optional[str] = None

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def validate_mobile(mobile: str) -> tuple[bool, str]:
    """Validate Indian mobile number."""
    if not mobile or len(mobile) != 10:
        return False, "Mobile number must be 10 digits"
    
    if not mobile.isdigit():
        return False, "Mobile number must contain only digits"
    
    if not mobile.startswith(('6', '7', '8', '9')):
        return False, "Invalid mobile number format"
    
    return True, "Valid"

def validate_gstin(gstin: str) -> tuple[bool, str]:
    """Validate GSTIN format."""
    if not gstin or len(gstin) != 15:
        return False, "GSTIN must be 15 characters"
    
    gstin = gstin.upper()
    if not re.match(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$', gstin):
        return False, "Invalid GSTIN format"
    
    return True, "Valid"

def generate_filename(prefix: str, extension: str) -> str:
    """Generate unique filename with timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{extension}"

# =============================================================================
# ERROR HANDLING DECORATOR
# =============================================================================

def handle_api_errors(func):
    """Decorator for handling API errors consistently"""
    @wraps(func)
    async def wrapper(*args, **kwargs):
        try:
            return await func(*args, **kwargs)
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"API error in {func.__name__}: {e}", exc_info=True)
            return JSONResponse(
                status_code=500,
                content={"success": False, "error": "Internal server error"}
            )
    return wrapper

# =============================================================================
# ENHANCED CACHE MANAGER
# =============================================================================

class CacheManager:
    def __init__(self, redis_url: Optional[str] = None):
        """Initialize cache manager with Redis fallback to in-memory."""
        self.redis_client = None
        self.memory_cache = {}
        self.cache_ttl = 300
        self._cleanup_task = None
        self._max_memory_cache_size = 1000
        
        if redis_url and HAS_REDIS:
            try:
                self.redis_client = redis.from_url(redis_url)
                logger.info("✅ Redis cache initialized")
            except Exception as e:
                logger.warning(f"Redis unavailable, using memory cache: {e}")
        
        if not self.redis_client:
            self._cleanup_task = asyncio.create_task(self._cleanup_memory_cache())
    
    async def get(self, key: str) -> Optional[Any]:
        """Get value from cache with proper cleanup."""
        if self.redis_client:
            try:
                data = await self.redis_client.get(key)
                return json.loads(data) if data else None
            except Exception as e:
                logger.warning(f"Redis get failed: {e}")
        
        # Memory cache fallback with size limits
        if len(self.memory_cache) > self._max_memory_cache_size:
            self._cleanup_expired_memory_cache()
        
        item = self.memory_cache.get(key)
        if item and item['expires'] > time.time():
            return item['data']
        elif item:
            del self.memory_cache[key]
        return None

    async def set(self, key: str, value: Any, ttl: int = None) -> bool:
        """Set cache value with TTL."""
        ttl = ttl or self.cache_ttl
        
        if self.redis_client:
            try:
                await self.redis_client.set(key, json.dumps(value), ex=ttl)
                return True
            except Exception as e:
                logger.warning(f"Redis set failed: {e}")
        
        # Memory cache fallback
        self.memory_cache[key] = {
            'data': value,
            'expires': time.time() + ttl
        }
        return True

    async def delete(self, key: str) -> bool:
        """Delete cache key."""
        if self.redis_client:
            try:
                await self.redis_client.delete(key)
            except Exception as e:
                logger.warning(f"Redis delete failed: {e}")
        
        self.memory_cache.pop(key, None)
        return True

    async def clear(self) -> bool:
        """Clear all cache."""
        if self.redis_client:
            try:
                await self.redis_client.flushdb()
            except Exception as e:
                logger.warning(f"Redis clear failed: {e}")
        
        self.memory_cache.clear()
        return True

    def _cleanup_expired_memory_cache(self):
        """Clean up expired entries and limit size."""
        current_time = time.time()
        expired_keys = [
            k for k, v in self.memory_cache.items() 
            if v['expires'] <= current_time
        ]
        for key in expired_keys:
            del self.memory_cache[key]
        
        # If still too large, remove oldest entries
        if len(self.memory_cache) > 800:
            sorted_items = sorted(
                self.memory_cache.items(), 
                key=lambda x: x[1]['expires']
            )
            for key, _ in sorted_items[:200]:  # Remove oldest 200
                del self.memory_cache[key]

    async def _cleanup_memory_cache(self):
        """Background task to cleanup memory cache."""
        while True:
            try:
                await asyncio.sleep(60)  # Run every minute
                self._cleanup_expired_memory_cache()
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Cache cleanup error: {e}")
                await asyncio.sleep(60)

    async def close(self):
        """Close cache connections."""
        if self._cleanup_task:
            self._cleanup_task.cancel()
            try:
                await self._cleanup_task
            except asyncio.CancelledError:
                pass
        
        if self.redis_client:
            await self.redis_client.close()
        
        self.memory_cache.clear()

# =============================================================================
# RATE LIMITER
# =============================================================================

class RateLimiter:
    def __init__(self, max_requests: int = config.RATE_LIMIT_REQUESTS, 
                 window_seconds: int = config.RATE_LIMIT_WINDOW):
        self.requests = defaultdict(list)
        self.max_requests = max_requests
        self.window = timedelta(seconds=window_seconds)
        self._last_cleanup = datetime.now()
        self._cleanup_interval = timedelta(hours=1)
    
    def is_allowed(self, identifier: str) -> bool:
        now = datetime.now()
        
        # Clean up old entries periodically to prevent memory leaks
        if now - self._last_cleanup > self._cleanup_interval:
            self._cleanup_old_entries(now)
            self._last_cleanup = now
        
        # Clean recent requests for this identifier
        self.requests[identifier] = [
            req_time for req_time in self.requests[identifier]
            if now - req_time < self.window
        ]
        
        if len(self.requests[identifier]) >= self.max_requests:
            return False
        
        self.requests[identifier].append(now)
        return True
    
    def _cleanup_old_entries(self, now: datetime):
        """Remove old entries to prevent memory leaks"""
        cutoff = now - self.window
        for identifier in list(self.requests.keys()):
            self.requests[identifier] = [
                req_time for req_time in self.requests[identifier]
                if req_time > cutoff
            ]
            if not self.requests[identifier]:
                del self.requests[identifier]

# =============================================================================
# GST API CLIENT
# =============================================================================

class GSTAPIClient:
    def __init__(self, api_key: str, host: str):
        self.api_key = api_key
        self.host = host
        self.base_url = f"https://{host}"
    
    async def fetch_gstin_data(self, gstin: str) -> dict:
        """Fetch GSTIN data from external API."""
        headers = {
            "X-RapidAPI-Key": self.api_key,
            "X-RapidAPI-Host": self.host
        }
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(
                f"{self.base_url}/gstin/{gstin}",
                headers=headers
            )
            
            if response.status_code == 200:
                data = response.json()
                # Normalize response
                return {
                    'gstin': gstin,
                    'company_name': data.get('tradeNam', 'Unknown'),
                    'status': data.get('sts', 'Unknown'),
                    'registration_date': data.get('rgdt'),
                    'constitution': data.get('ctb'),
                    'compliance_score': self._calculate_compliance_score(data),
                    'raw_data': data
                }
            else:
                raise Exception(f"API request failed: {response.status_code}")
    
    def _calculate_compliance_score(self, data: dict) -> int:
        """Calculate compliance score based on GST data."""
        score = 50  # Base score
        
        if data.get('sts') == 'Active':
            score += 30
        
        if data.get('rgdt'):
            # Bonus for older registrations
            try:
                reg_date = datetime.strptime(data['rgdt'], '%Y-%m-%d')
                years_active = (datetime.now() - reg_date).days / 365
                score += min(20, int(years_active * 2))
            except:
                pass
        
        return min(100, max(0, score))

# =============================================================================
# INITIALIZE FASTAPI APP
# =============================================================================

app = FastAPI(
    title="GST Intelligence Platform", 
    version="2.0.0",
    description="Advanced GST Compliance Analytics Platform with Loan Management"
)

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Initialize components
db = DatabaseManager()
cache_manager = CacheManager(config.REDIS_URL)
login_limiter = RateLimiter(max_requests=5, window_seconds=900)
api_limiter = RateLimiter()

# Initialize API client
api_client = GSTAPIClient(config.RAPIDAPI_KEY, config.RAPIDAPI_HOST) if config.RAPIDAPI_KEY else None

# Initialize loan manager
loan_manager = None
if HAS_LOAN_MANAGEMENT and config.RAZORPAY_KEY_ID and config.RAZORPAY_KEY_SECRET:
    try:
        razorpay_client = RazorpayLendingClient(
            key_id=config.RAZORPAY_KEY_ID,
            key_secret=config.RAZORPAY_KEY_SECRET,
            environment=config.RAZORPAY_ENVIRONMENT
        )
        loan_manager = LoanManager(razorpay_client, db)
        logger.info("✅ Loan management system initialized")
    except Exception as e:
        logger.error(f"Failed to initialize loan management: {e}")

# Initialize AI client
ai_client = None
if HAS_AI_FEATURES and config.ANTHROPIC_API_KEY:
    try:
        ai_client = AnthropicClient(config.ANTHROPIC_API_KEY)
        logger.info("✅ AI client initialized")
    except Exception as e:
        logger.error(f"Failed to initialize AI client: {e}")

# =============================================================================
# AUTHENTICATION HELPERS
# =============================================================================

async def get_current_user(request: Request) -> Optional[str]:
    """Get current user from session."""
    session_token = request.cookies.get("session_token")
    if not session_token:
        return None
    return await db.get_session(session_token)

async def get_current_user_or_raise(request: Request, admin_required: bool = False) -> str:
    """Unified auth check with optional admin requirement"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            headers={"Location": "/login"}
        )
    
    if admin_required:
        admin_users = config.ADMIN_USERS.split(",") if config.ADMIN_USERS else []
        if user not in admin_users:
            raise HTTPException(status_code=403, detail="Admin access required")
    
    return user

async def require_auth(request: Request) -> str:
    """Require authentication for protected routes."""
    return await get_current_user_or_raise(request)

async def require_admin(request: Request) -> str:
    """Require admin authentication."""
    return await get_current_user_or_raise(request, admin_required=True)

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

async def fetch_company_data(gstin: str) -> Optional[dict]:
    """Fetch and cache company data."""
    cache_key = f"gstin:{gstin}"
    
    cached_data = await cache_manager.get(cache_key)
    if cached_data:
        logger.info(f"Cache hit for GSTIN: {gstin}")
        return cached_data
    
    if api_client:
        try:
            company_data = await api_client.fetch_gstin_data(gstin)
            await cache_manager.set(cache_key, company_data, ttl=3600)
            logger.info(f"Cached GSTIN data: {gstin}")
            return company_data
        except Exception as e:
            logger.error(f"API fetch failed for {gstin}: {e}")
            raise
    
    return None

async def get_user_display_name(mobile: str) -> str:
    """Get user display name or fallback to mobile."""
    try:
        profile = await db.get_user_profile(mobile)
        display_name = profile.get('display_name')
        if display_name:
            return display_name
        return f"User {mobile[-4:]}"
    except:
        return f"User {mobile[-4:]}"

# =============================================================================
# MAIN ROUTES
# =============================================================================

@app.get("/health")
async def health_check():
    """Health check endpoint."""
    try:
        # Test database connection
        if db.pool:
            async with db.pool.acquire() as conn:
                await conn.execute("SELECT 1")
        
        return {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "version": "2.0.0",
            "checks": {
                "database": "healthy",
                "gst_api": "configured" if config.RAPIDAPI_KEY else "missing",
                "ai_features": "configured" if config.ANTHROPIC_API_KEY else "disabled",
                "cache": "redis" if cache_manager.redis_client else "memory",
                "loan_management": "available" if HAS_LOAN_MANAGEMENT else "disabled",
                "pdf_generation": "available" if HAS_WEASYPRINT else "disabled"
            }
        }
    except Exception as e:
        return JSONResponse(
            status_code=503,
            content={
                "status": "unhealthy", 
                "error": str(e), 
                "timestamp": datetime.now().isoformat()
            }
        )

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, current_user: str = Depends(require_auth)):
    """Main dashboard page."""
    user_profile = await db.get_user_profile(current_user)
    user_display_name = await get_user_display_name(current_user)
    stats = await db.get_user_stats(current_user)
    recent_searches = await db.get_search_history(current_user, limit=5)
    
    return templates.TemplateResponse("index.html", {
        "request": request,
        "current_user": current_user,
        "user_display_name": user_display_name,
        "user_profile": user_profile,
        "stats": stats,
        "recent_searches": recent_searches
    })

# =============================================================================
# AUTHENTICATION ROUTES
# =============================================================================

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Login page."""
    # Check if already logged in
    user = await get_current_user(request)
    if user:
        return RedirectResponse(url="/", status_code=302)
    
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
async def login(request: Request, mobile: str = Form(...), password: str = Form(...)):
    """Handle login."""
    client_ip = request.client.host
    
    if not login_limiter.is_allowed(client_ip):
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Too many login attempts. Please try again later."
        })
    
    is_valid, message = validate_mobile(mobile)
    if not is_valid:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": message
        })
    
    if await db.verify_user(mobile, password):
        session_token = await db.create_session(mobile)
        response = RedirectResponse(url="/", status_code=302)
        response.set_cookie(
            "session_token", 
            session_token, 
            max_age=int(config.SESSION_DURATION.total_seconds()),
            httponly=True,
            secure=config.is_production
        )
        return response
    else:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Invalid mobile number or password"
        })

@app.get("/signup", response_class=HTMLResponse)
async def signup_page(request: Request):
    """Signup page."""
    return templates.TemplateResponse("signup.html", {"request": request})

@app.post("/signup")
async def signup(request: Request, mobile: str = Form(...), 
                password: str = Form(...), confirm_password: str = Form(...)):
    """Handle signup."""
    is_valid, message = validate_mobile(mobile)
    if not is_valid:
        return templates.TemplateResponse("signup.html", {
            "request": request,
            "error": message
        })
    
    if len(password) < 6:
        return templates.TemplateResponse("signup.html", {
            "request": request,
            "error": "Password must be at least 6 characters"
        })
    
    if password != confirm_password:
        return templates.TemplateResponse("signup.html", {
            "request": request,
            "error": "Passwords do not match"
        })
    
    salt = secrets.token_hex(16)
    password_hash = hashlib.pbkdf2_hmac(
        'sha256', password.encode('utf-8'), 
        salt.encode('utf-8'), 100000, dklen=64
    ).hex()
    
    success = await db.create_user(mobile, password_hash, salt)
    if not success:
        return templates.TemplateResponse("signup.html", {
            "request": request,
            "error": "Mobile number already registered"
        })
    
    return RedirectResponse(url="/login?registered=true", status_code=302)

@app.get("/logout")
async def logout(request: Request):
    """Handle logout."""
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.delete_session(session_token)
    
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("session_token")
    return response

# =============================================================================
# SEARCH ROUTES
# =============================================================================

@app.get("/search")
async def search_gstin_get(request: Request, gstin: str = None, current_user: str = Depends(require_auth)):
    """Handle GET search requests."""
    if not gstin:
        return RedirectResponse(url="/", status_code=302)
    
    return await process_search(request, gstin, current_user)

@app.post("/search")
async def search_gstin_post(request: Request, gstin: str = Form(...), current_user: str = Depends(require_auth)):
    """Handle POST search requests."""
    return await process_search(request, gstin, current_user)

async def process_search(request: Request, gstin: str, current_user: str):
    """Process GST search request."""
    user_profile = await db.get_user_profile(current_user)
    user_display_name = await get_user_display_name(current_user)
    
    # Validate GSTIN
    is_valid, message = validate_gstin(gstin.upper())
    if not is_valid:
        return templates.TemplateResponse("search_results.html", {
            "request": request,
            "current_user": current_user,
            "user_display_name": user_display_name,
            "user_profile": user_profile,
            "error": message,
            "gstin": gstin
        })
    
    try:
        # Fetch company data
        company_data = await fetch_company_data(gstin.upper())
        
        if company_data:
            # Save to search history
            await db.save_search(current_user, gstin.upper(), company_data)
            
            # Get related searches
            history = await db.get_search_history(current_user, limit=5)
            
            return templates.TemplateResponse("search_results.html", {
                "request": request,
                "current_user": current_user,
                "user_display_name": user_display_name,
                "user_profile": user_profile,
                "company_data": company_data,
                "history": history
            })
        else:
            return templates.TemplateResponse("search_results.html", {
                "request": request,
                "current_user": current_user,
                "user_display_name": user_display_name,
                "user_profile": user_profile,
                "error": "Company data not found or API unavailable",
                "gstin": gstin
            })
            
    except Exception as e:
        logger.error(f"Search error for {gstin}: {e}")
        return templates.TemplateResponse("search_results.html", {
            "request": request,
            "current_user": current_user,
            "user_display_name": user_display_name,
            "user_profile": user_profile,
            "error": "Search failed. Please try again.",
            "gstin": gstin
        })

# =============================================================================
# BATCH SEARCH ROUTES
# =============================================================================

@app.post("/api/search/batch")
@handle_api_errors
async def batch_search_api(request: BatchSearchRequest, current_user: str = Depends(require_auth)):
    """Batch GSTIN search API."""
    results = []
    errors = []
    
    for gstin in request.gstins:
        try:
            company_data = await fetch_company_data(gstin)
            if company_data:
                await db.save_search(current_user, gstin, company_data)
                results.append(company_data)
            else:
                errors.append(f"No data found for {gstin}")
        except Exception as e:
            logger.error(f"Batch search error for {gstin}: {e}")
            errors.append(f"Error searching {gstin}: {str(e)}")
    
    return {
        "success": True,
        "results": results,
        "errors": errors,
        "total_searched": len(request.gstins),
        "successful": len(results),
        "failed": len(errors)
    }

# =============================================================================
# PDF GENERATION ROUTES
# =============================================================================

@app.get("/generate-pdf")
async def generate_pdf_report(request: Request, gstin: str, current_user: str = Depends(require_auth)):
    """Generate PDF report for GSTIN."""
    if not HAS_WEASYPRINT:
        raise HTTPException(status_code=503, detail="PDF generation not available")
    
    try:
        # Get company data
        company_data = await fetch_company_data(gstin.upper())
        if not company_data:
            raise HTTPException(status_code=404, detail="Company data not found")
        
        # Get user info
        user_profile = await db.get_user_profile(current_user)
        user_display_name = await get_user_display_name(current_user)
        
        # Render HTML template
        html_content = templates.get_template("pdf_report.html").render({
            "company_data": company_data,
            "user_profile": user_profile,
            "user_display_name": user_display_name,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "request": request
        })
        
        # Generate PDF
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        # Create response
        filename = f"GST_Report_{gstin}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"PDF generation error for {gstin}: {e}")
        raise HTTPException(status_code=500, detail="PDF generation failed")

# =============================================================================
# EXPORT ROUTES
# =============================================================================

@app.post("/api/user/export")
@handle_api_errors
async def export_user_data(request: ExportRequest, current_user: str = Depends(require_auth)):
    """Export user data in various formats."""
    try:
        # Get data based on type
        if request.type == "history":
            data = await db.get_all_searches(current_user)
            filename_prefix = "search_history"
        elif request.type == "analytics":
            data = await db.get_analytics_data(current_user)
            filename_prefix = "analytics_data"
        elif request.type == "all":
            # Combine all data
            history = await db.get_all_searches(current_user)
            analytics = await db.get_analytics_data(current_user)
            profile = await db.get_user_profile(current_user)
            data = {
                "search_history": history,
                "analytics": analytics,
                "profile": profile
            }
            filename_prefix = "complete_data"
        else:
            raise HTTPException(status_code=400, detail="Invalid export type")
        
        # Generate export based on format
        if request.format == "csv":
            return await export_as_csv(data, filename_prefix, request.type)
        elif request.format == "excel":
            return await export_as_excel(data, filename_prefix, request.type)
        elif request.format == "pdf":
            return await export_as_pdf(data, filename_prefix, request.type, current_user)
        else:
            raise HTTPException(status_code=400, detail="Invalid export format")
            
    except Exception as e:
        logger.error(f"Export error: {e}")
        raise HTTPException(status_code=500, detail="Export failed")

async def export_as_csv(data: Union[List, Dict], filename_prefix: str, export_type: str) -> StreamingResponse:
    """Export data as CSV."""
    output = StringIO()
    
    if export_type == "history" and isinstance(data, list):
        writer = csv.DictWriter(output, fieldnames=[
            'gstin', 'company_name', 'status', 'compliance_score', 
            'searched_at', 'search_count'
        ])
        writer.writeheader()
        for item in data:
            writer.writerow({
                'gstin': item.get('gstin', ''),
                'company_name': item.get('company_name', ''),
                'status': item.get('status', ''),
                'compliance_score': item.get('compliance_score', ''),
                'searched_at': item.get('searched_at', ''),
                'search_count': item.get('search_count', '')
            })
    
    output.seek(0)
    filename = generate_filename(filename_prefix, "csv")
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

async def export_as_excel(data: Union[List, Dict], filename_prefix: str, export_type: str) -> StreamingResponse:
    """Export data as Excel file."""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Export"
        
        if export_type == "history" and isinstance(data, list):
            # Headers
            headers = ['GSTIN', 'Company Name', 'Status', 'Compliance Score', 'Searched At', 'Search Count']
            ws.append(headers)
            
            # Data rows
            for item in data:
                ws.append([
                    item.get('gstin', ''),
                    item.get('company_name', ''),
                    item.get('status', ''),
                    item.get('compliance_score', ''),
                    str(item.get('searched_at', '')),
                    item.get('search_count', '')
                ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = generate_filename(filename_prefix, "xlsx")
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(status_code=503, detail="Excel export not available")

async def export_as_pdf(data: Union[List, Dict], filename_prefix: str, export_type: str, current_user: str) -> StreamingResponse:
    """Export data as PDF."""
    if not HAS_WEASYPRINT:
        raise HTTPException(status_code=503, detail="PDF export not available")
    
    try:
        user_profile = await db.get_user_profile(current_user)
        
        # Render HTML template for export
        html_content = templates.get_template("export_pdf.html").render({
            "data": data,
            "export_type": export_type,
            "user_profile": user_profile,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        
        # Generate PDF
        pdf_bytes = HTML(string=html_content).write_pdf()
        
        filename = generate_filename(filename_prefix, "pdf")
        
        return StreamingResponse(
            BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        raise HTTPException(status_code=500, detail="PDF export failed")

@app.get("/export/history")
async def export_search_history(request: Request, format: str = "csv", current_user: str = Depends(require_auth)):
    """Quick export search history."""
    export_request = ExportRequest(format=format, type="history")
    return await export_user_data(export_request, current_user)

# =============================================================================
# FILE UPLOAD ROUTES
# =============================================================================

@app.post("/upload/batch-search")
async def upload_batch_search(
    request: Request,
    file: UploadFile = File(...),
    current_user: str = Depends(require_auth)
):
    """Upload CSV/Excel file for batch GSTIN search."""
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Use CSV or Excel files.")
    
    try:
        contents = await file.read()
        
        # Parse file based on type
        if file.filename.endswith('.csv'):
            gstins = await parse_csv_for_gstins(contents)
        else:
            gstins = await parse_excel_for_gstins(contents)
        
        if not gstins:
            raise HTTPException(status_code=400, detail="No valid GSTINs found in file")
        
        if len(gstins) > 50:  # Limit for bulk processing
            raise HTTPException(status_code=400, detail="Maximum 50 GSTINs allowed per upload")
        
        # Process batch search
        batch_request = BatchSearchRequest(gstins=gstins)
        return await batch_search_api(batch_request, current_user)
        
    except Exception as e:
        logger.error(f"Batch upload error: {e}")
        raise HTTPException(status_code=500, detail="File processing failed")

async def parse_csv_for_gstins(contents: bytes) -> List[str]:
    """Parse CSV file to extract GSTINs."""
    try:
        csv_text = contents.decode('utf-8')
        reader = csv.DictReader(StringIO(csv_text))
        
        gstins = []
        for row in reader:
            # Look for GSTIN in common column names
            gstin = None
            for key in row.keys():
                if key.lower() in ['gstin', 'gst', 'tin', 'gstin_number']:
                    gstin = row[key].strip().upper()
                    break
            
            if gstin and validate_gstin(gstin)[0]:
                gstins.append(gstin)
        
        return list(set(gstins))  # Remove duplicates
        
    except Exception as e:
        logger.error(f"CSV parsing error: {e}")
        return []

async def parse_excel_for_gstins(contents: bytes) -> List[str]:
    """Parse Excel file to extract GSTINs."""
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        gstins = []
        header_row = None
        gstin_col = None
        
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if header_row is None:
                # Find GSTIN column
                for col_num, cell_value in enumerate(row):
                    if cell_value and str(cell_value).lower() in ['gstin', 'gst', 'tin', 'gstin_number']:
                        gstin_col = col_num
                        header_row = row_num
                        break
            else:
                # Extract GSTIN from the identified column
                if gstin_col is not None and len(row) > gstin_col:
                    gstin = str(row[gstin_col]).strip().upper() if row[gstin_col] else None
                    if gstin and validate_gstin(gstin)[0]:
                        gstins.append(gstin)
        
        return list(set(gstins))  # Remove duplicates
        
    except ImportError:
        raise HTTPException(status_code=503, detail="Excel processing not available")
    except Exception as e:
        logger.error(f"Excel parsing error: {e}")
        return []

# =============================================================================
# PAGE ROUTES
# =============================================================================

@app.get("/history", response_class=HTMLResponse)
async def history_page(request: Request, current_user: str = Depends(require_auth)):
    """Search history page."""
    history = await db.get_all_searches(current_user)
    user_profile = await db.get_user_profile(current_user)
    user_display_name = await get_user_display_name(current_user)
    
    now = datetime.now()
    last_30_days = now - timedelta(days=30)
    searches_this_month = sum(
        1 for item in history
        if item.get('searched_at') and item['searched_at'] >= last_30_days
    )
    
    return templates.TemplateResponse("history.html", {
        "request": request, 
        "current_user": current_user,
        "user_display_name": user_display_name,
        "user_profile": user_profile,
        "history": history,
        "searches_this_month": searches_this_month
    })

@app.get("/analytics", response_class=HTMLResponse)
async def analytics_page(request: Request, current_user: str = Depends(require_auth)):
    """Analytics dashboard page."""
    user_profile = await db.get_user_profile(current_user)
    user_display_name = await get_user_display_name(current_user)
    analytics_data = await db.get_analytics_data(current_user)
    
    stats = await db.get_user_stats(current_user)
    
    return templates.TemplateResponse("analytics.html", {
        "request": request, 
        "current_user": current_user,
        "user_display_name": user_display_name,
        "user_profile": user_profile,
        "daily_searches": analytics_data["daily_searches"],
        "score_distribution": analytics_data["score_distribution"],
        "top_companies": analytics_data["top_companies"],
        "total_searches": stats["total_searches"],
        "unique_companies": stats["unique_companies"],
        "avg_compliance": stats["avg_compliance"]
    })

# =============================================================================
# LOAN ROUTES
# =============================================================================

@app.get("/loans", response_class=HTMLResponse)
async def loans_page(request: Request, current_user: str = Depends(require_auth)):
    """Loans dashboard page"""
    try:
        applications = []
        if loan_manager:
            applications = await loan_manager.get_user_loan_applications(current_user)
        
        user_profile = await db.get_user_profile(current_user)
        user_display_name = await get_user_display_name(current_user)
        
        loan_config_data = {}
        if HAS_LOAN_MANAGEMENT and LoanConfig:
            loan_config_data = {
                "min_amount": LoanConfig.MIN_LOAN_AMOUNT,
                "max_amount": LoanConfig.MAX_LOAN_AMOUNT,
                "min_compliance": LoanConfig.MIN_COMPLIANCE_SCORE,
                "min_vintage": LoanConfig.MIN_BUSINESS_VINTAGE_MONTHS
            }
        
        return templates.TemplateResponse("loans.html", {
            "request": request,
            "current_user": current_user,
            "user_display_name": user_display_name,
            "user_profile": user_profile,
            "applications": applications,
            "loan_config": loan_config_data
        })
    except Exception as e:
        logger.error(f"Loans page error: {e}")
        return templates.TemplateResponse("loans.html", {
            "request": request,
            "current_user": current_user,
            "error": "Failed to load loan data"
        })

@app.post("/api/loans/eligibility")
@handle_api_errors
async def check_loan_eligibility(request: Request, gstin: str = Form(...), current_user: str = Depends(require_auth)):
    """Check loan eligibility for a GSTIN."""
    if not HAS_LOAN_MANAGEMENT:
        raise HTTPException(status_code=503, detail="Loan management not available")
    
    try:
        # Get company data
        company_data = await fetch_company_data(gstin.upper())
        if not company_data:
            raise HTTPException(status_code=404, detail="Company data not found")
        
        # Check eligibility using loan manager
        eligibility = await loan_manager.check_eligibility(current_user, company_data)
        
        return {"success": True, "data": eligibility}
        
    except Exception as e:
        logger.error(f"Eligibility check error: {e}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": "Failed to check eligibility"}
        )

# =============================================================================
# API ROUTES
# =============================================================================

@app.get("/api/user/stats")
@handle_api_errors
async def get_user_stats_api(current_user: str = Depends(require_auth)):
    """Get user statistics API."""
    stats = await db.get_user_stats(current_user)
    return {"success": True, "data": stats}

@app.post("/api/user/change-password")
@handle_api_errors
async def change_password_api(request: ChangePasswordRequest, current_user: str = Depends(require_auth)):
    """Change user password API."""
    success = await db.change_password(current_user, request.currentPassword, request.newPassword)
    
    if success:
        return {"success": True, "message": "Password changed successfully"}
    else:
        return {"success": False, "message": "Current password is incorrect"}

@app.get("/api/user/preferences")
@handle_api_errors
async def get_user_preferences_api(current_user: str = Depends(require_auth)):
    """Get user preferences API."""
    preferences = await db.get_user_preferences(current_user)
    return {"success": True, "data": preferences}

@app.post("/api/user/preferences")
@handle_api_errors
async def save_user_preferences_api(request: UserPreferencesRequest, current_user: str = Depends(require_auth)):
    """Save user preferences API."""
    preferences = {k: v for k, v in request.dict().items() if v is not None}
    await db.save_user_preferences(current_user, preferences)
    return {"success": True, "message": "Preferences saved successfully"}

@app.get("/api/user/profile")
@handle_api_errors
async def get_user_profile_api(current_user: str = Depends(require_auth)):
    """Get user profile API."""
    profile = await db.get_user_profile(current_user)
    return {"success": True, "data": profile}

@app.post("/api/user/profile")
@handle_api_errors
async def save_user_profile_api(request: UserProfileRequest, current_user: str = Depends(require_auth)):
    """Save user profile API."""
    profile_data = {k: v for k, v in request.dict().items() if v is not None}
    await db.save_user_profile(current_user, profile_data)
    return {"success": True, "message": "Profile saved successfully"}

@app.delete("/api/user/history/clear")
@handle_api_errors
async def clear_search_history_api(current_user: str = Depends(require_auth)):
    """Clear user search history."""
    success = await db.clear_search_history(current_user)
    if success:
        return {"success": True, "message": "Search history cleared"}
    else:
        return {"success": False, "message": "Failed to clear history"}

@app.delete("/api/user/history/{gstin}")
@handle_api_errors
async def delete_search_item_api(gstin: str, current_user: str = Depends(require_auth)):
    """Delete specific search from history."""
    success = await db.delete_search(current_user, gstin.upper())
    if success:
        return {"success": True, "message": "Search deleted"}
    else:
        return {"success": False, "message": "Failed to delete search"}

# =============================================================================
# ERROR LOGGING API
# =============================================================================

@app.post("/api/system/error")
async def log_client_error(
    request: Request,
    error_data: ErrorLogRequest,
    current_user: Optional[str] = Depends(get_current_user)
):
    """Log client-side errors."""
    try:
        # Convert to dict and add user context
        log_data = error_data.dict()
        log_data['mobile'] = current_user
        log_data['ip_address'] = request.client.host
        
        # Log to database
        await db.log_error(log_data)
        
        return {"success": True, "message": "Error logged"}
    except Exception as e:
        logger.error(f"Failed to log client error: {e}")
        return {"success": False, "error": "Failed to log error"}

# =============================================================================
# ADMIN ROUTES
# =============================================================================

@app.get("/admin", response_class=HTMLResponse)
async def admin_dashboard(request: Request, current_user: str = Depends(require_admin)):
    """Admin dashboard page."""
    user_profile = await db.get_user_profile(current_user)
    user_display_name = await get_user_display_name(current_user)
    system_stats = await db.get_system_stats()
    
    return templates.TemplateResponse("admin_dashboard.html", {
        "request": request,
        "current_user": current_user,
        "user_display_name": user_display_name,
        "user_profile": user_profile,
        "system_stats": system_stats
    })

@app.post("/api/admin/bulk/delete-old-searches")
@handle_api_errors
async def bulk_delete_old_searches(
    request: Request,
    days_old: int = 90,
    current_user: str = Depends(require_admin)
):
    """Delete search history older than specified days with proper validation."""
    # Input validation
    if not isinstance(days_old, int) or days_old < 1 or days_old > 3650:
        raise HTTPException(
            status_code=400, 
            detail="Invalid days_old parameter. Must be between 1 and 3650."
        )
    
    try:
        async with db.pool.acquire() as conn:
            result = await conn.fetchval("""
                WITH deleted AS (
                    DELETE FROM search_history 
                    WHERE searched_at < CURRENT_DATE - INTERVAL '%s days'
                    RETURNING id
                )
                SELECT COUNT(*) FROM deleted
            """, days_old)
            
            # Log admin action
            logger.info(f"Admin {current_user} deleted {result or 0} search records older than {days_old} days")
            
            return {
                'success': True,
                'deleted_count': result or 0,
                'days_old': days_old,
                'message': f"Successfully deleted {result or 0} old search records"
            }
    except Exception as e:
        logger.error(f"Bulk delete error: {e}")
        return {'success': False, 'error': 'Failed to delete old searches'}

@app.get("/api/admin/users")
@handle_api_errors
async def get_all_users_admin(
    page: int = 1,
    per_page: int = 50,
    search: str = None,
    current_user: str = Depends(require_admin)
):
    """Get paginated list of all users (admin only)."""
    users_data = await db.get_all_users(page, per_page, search)
    return {"success": True, "data": users_data}

# =============================================================================
# STATIC FILES
# =============================================================================

@app.get("/favicon.ico")
async def favicon():
    """Favicon route."""
    favicon_bytes = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\rIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00\x05\x00\x01\x8d\xc8\x02\x00\x00\x00\x00IEND\xaeB`\x82'
    return Response(content=favicon_bytes, media_type="image/png")

@app.get("/manifest.json")
async def get_manifest():
    """PWA manifest route."""
    with open("static/manifest.json", "r") as f:
        manifest_content = f.read()
    return Response(content=manifest_content, media_type="application/json")

@app.get("/sw.js")
async def get_service_worker():
    """Service worker route."""
    with open("sw.js", "r") as f:
        sw_content = f.read()
    return Response(content=sw_content, media_type="application/javascript")

# =============================================================================
# ERROR HANDLERS
# =============================================================================

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """Handle HTTP exceptions."""
    if exc.status_code == 303 and exc.headers and "Location" in exc.headers:
        return RedirectResponse(url=exc.headers["Location"], status_code=303)
    
    if exc.status_code == 404:
        return templates.TemplateResponse("errors/404.html", {"request": request}, status_code=404)
    
    return JSONResponse(
        status_code=exc.status_code, 
        content={"error": exc.detail}
    )

@app.exception_handler(500)
async def internal_server_error_handler(request: Request, exc: Exception):
    """Handle 500 errors with proper error page."""
    logger.error(f"Internal server error: {exc}", exc_info=True)
    
    # Log error to database if available
    try:
        current_user = await get_current_user(request)
        await db.log_error({
            'type': 'Internal Server Error',
            'message': str(exc),
            'stack': str(exc.__traceback__) if exc.__traceback__ else None,
            'url': str(request.url),
            'userAgent': request.headers.get('user-agent'),
            'mobile': current_user
        })
    except:
        pass  # Don't fail on error logging
    
    # Return error page
    return templates.TemplateResponse(
        "errors/500.html", 
        {"request": request, "error": str(exc)},
        status_code=500
    )

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    """Handle general exceptions."""
    logger.error(f"Unhandled exception: {exc}", exc_info=True)
    return JSONResponse(
        status_code=500, 
        content={"error": "An internal error occurred"}
    )

# =============================================================================
# STARTUP/SHUTDOWN EVENTS
# =============================================================================

@app.on_event("startup")
async def startup():
    """Initialize application on startup with better error handling."""
    logger.info("🚀 GST Intelligence Platform starting up...")
    
    try:
        # Initialize database with retry
        await db.initialize()
        await db.cleanup_expired_sessions()
        logger.info("✅ Database initialized")
        
        # Initialize loan management if available
        if loan_manager:
            logger.info("✅ Loan management system ready")
        
        # Set admin users
        admin_users = config.ADMIN_USERS.split(",") if config.ADMIN_USERS else []
        app.extra = {"ADMIN_USERS": admin_users}
        logger.info(f"✅ Admin users configured: {len(admin_users)} users")
        
        # Test external APIs
        if config.RAPIDAPI_KEY:
            logger.info("✅ GST API configured")
        else:
            logger.warning("⚠️ GST API not configured - search functionality limited")
        
        if config.ANTHROPIC_API_KEY:
            logger.info("✅ AI features enabled")
        else:
            logger.info("ℹ️ AI features disabled")
        
        if HAS_WEASYPRINT:
            logger.info("✅ PDF generation enabled")
        else:
            logger.warning("⚠️ PDF generation disabled")
        
        logger.info("🎉 Application startup complete")
        
    except Exception as e:
        logger.error(f"❌ Startup failed: {e}")
        raise

@app.on_event("shutdown")
async def shutdown():
    """Clean up on shutdown."""
    logger.info("🔄 GST Intelligence Platform shutting down...")
    
    try:
        # Close cache manager
        await cache_manager.close()
        logger.info("✅ Cache manager closed")
        
        # Close database pool
        if db.pool:
            await db.pool.close()
            logger.info("✅ Database pool closed")
        
        logger.info("✅ Cleanup complete")
        
    except Exception as e:
        logger.error(f"Error during shutdown: {e}")

# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        app, 
        host="0.0.0.0", 
        port=int(os.getenv("PORT", "8000")),
        workers=int(os.getenv("WORKERS", "1"))
    )